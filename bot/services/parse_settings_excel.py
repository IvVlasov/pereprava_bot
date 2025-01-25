import logging

import openpyxl
from aiogram.types.input_file import FSInputFile
from openpyxl.worksheet.worksheet import Worksheet

from bot.app import bot
from bot.constants.buttons import ModeratorMessageTypes
from bot.models.crossing import Crossing
from bot.models.manager_link import ManagerLink
from bot.models.message_template import MessageTemplate, MessageTemplateType
from repository import CrossingRepository, MessageTemplatesRepository
from settings import get_settings
from repository.messages import MessagesRepository
from bot.models.messages import Message

logger = logging.getLogger(__name__)


class ExcelSettings:
    def __init__(self):
        settings = get_settings()
        self.file_path = settings.SETTINGS_FILE_PATH
        self.crossing_repository = CrossingRepository()
        self.message_templates_repository = MessageTemplatesRepository()

    async def parse_and_save(self) -> bool:
        # try:
        wb_obj = openpyxl.load_workbook(self.file_path)
        for sheet_name in wb_obj.sheetnames:
            sheet_obj = wb_obj[sheet_name]
            print(sheet_name)
            if sheet_name.strip() == "Переправы":
                await self.save_crossings(sheet_obj)
            elif sheet_name.strip() == "Утреннее сообщение":
                await self.save_message_templates(sheet_obj)
            elif sheet_name.strip() == "Статические сообщения":
                await self.save_static_message_templates(sheet_obj)
            elif sheet_name.strip() == "Сообщения":
                await self.save_messages(sheet_obj)
        return True
        # except Exception as e:
        #     logger.error(f"Error parsing settings excel: {e}")
        #     return False

    async def save_messages(self, sheet_obj: Worksheet):
        messages_repository = MessagesRepository()
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            message = Message(key=row[0], name=row[1], text=row[2])
            await messages_repository.create_message(message)

    def _get_message_template_type(
        self, message_template_name: str
    ) -> MessageTemplateType:
        match message_template_name:
            case ModeratorMessageTypes.MORNING_MESSAGE:
                return MessageTemplateType.MORNING
            case ModeratorMessageTypes.CLOSING_MESSAGE:
                return MessageTemplateType.CLOSING
            case ModeratorMessageTypes.OPENING_MESSAGE:
                return MessageTemplateType.OPENING
            case ModeratorMessageTypes.PASSENGER_TRAIN:
                return MessageTemplateType.PASSENGER_TRAIN
            case ModeratorMessageTypes.LIMIT_MESSAGE:
                return MessageTemplateType.LIMIT
            case _:
                raise ValueError(
                    f"Invalid message template type: {message_template_name}"
                )

    async def save_static_message_templates(self, sheet_obj: Worksheet):
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            message_template = MessageTemplate(
                name=row[0],
                template=row[1],
                message_type=self._get_message_template_type(row[0]),
            )
            await self.message_templates_repository.create_or_update_message_template(
                message_template
            )

    async def save_crossings(self, sheet_obj: Worksheet):
        current_crossings = await self.crossing_repository.get_all_crossings()
        current_crossings_names = [crossing.name for crossing in current_crossings]
        config_crossings_names = []
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            crossing = Crossing(name=row[0], camera_url=row[1])
            config_crossings_names.append(crossing.name)
            await self.crossing_repository.create_or_update_crossing(crossing)

        for current_crossing_name in current_crossings_names:
            if current_crossing_name not in config_crossings_names:
                await self.crossing_repository.delete_crossing(current_crossing_name)

    async def save_message_templates(self, sheet_obj: Worksheet):
        current_message_templates = (
            await self.message_templates_repository.get_all_message_templates()
        )
        current_message_templates_names = [
            message_template.name for message_template in current_message_templates
        ]
        config_message_templates_names = []
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            message_template = MessageTemplate(
                name=row[0], template=row[1], message_type=MessageTemplateType.MORNING
            )
            config_message_templates_names.append(message_template.name)
            await self.message_templates_repository.create_or_update_message_template(
                message_template
            )

        for current_message_template_name in current_message_templates_names:
            if current_message_template_name not in config_message_templates_names:
                await self.message_templates_repository.delete_message_template(
                    current_message_template_name
                )

    async def _get_link(self, manager_link: ManagerLink):
        tg_bot = await bot.get_me()
        return f"https://t.me/{tg_bot.username}?start={manager_link.uuid}"

    async def get_excel_file(self):
        return FSInputFile(self.file_path)
